#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2022-01-09 20:41
# @Author   :wangqinghua
# @Email    : 867075698@qq.com
# @File     : do_excel.py
# @Software : PyCharm

from openpyxl import load_workbook
from tools.read_config import Read_Config
from tools import project_path
import time
#绝对路径 相对路径

class Do_Excel():
    @staticmethod
    def get_data(file_name):
        wb=load_workbook(file_name)

        mode = eval(Read_Config().read_config(project_path.case_config_path, "MODE","mode"))


        tel = "数学课程"+str((int(time.time()*1000)))      #假设是从Eecel里面拿到的
        test_data = []
        for key in mode:
            sheet=wb[key]
            if mode[key]=="all":
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data["case_id"]=sheet.cell(i,1).value
                    row_data["module"] = sheet.cell(i, 2).value
                    row_data["title"] = sheet.cell(i, 3).value
                    row_data["url"]=sheet.cell(i,4).value
                    # row_data["data"] = sheet.cell(i, 5).value
                    if sheet.cell(i,5).value.find("tel_1")!=-1:
                        row_data['data']=sheet.cell(i,5).value.replace("tel_1",tel)
                    elif sheet.cell(i + 1, 5).value.find("tel")!= -1:
                        row_data['data'] = sheet.cell(i + 1, 5).value.replace("tel", tel+str(int(1)))
                    else:
                        row_data["data"] = sheet.cell(i, 5).value
                    row_data["expected"] = sheet.cell(i, 6).value
                    row_data["method"] = sheet.cell(i, 7).value
                    row_data["sheet_name"]=key  #加一个表单名
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data={}
                    row_data["case_id"]=sheet.cell(case_id+1,1).value
                    row_data["module"] = sheet.cell(case_id+1, 2).value
                    row_data["title"] = sheet.cell(case_id+1, 3).value
                    row_data["url"]=sheet.cell(case_id+1,4).value
                    # row_data["data"] = sheet.cell(case_id+1, 5).value
                    if sheet.cell(case_id+1,5).value.find("tel_1")!=-1:
                        row_data['data'] = sheet.cell(case_id+1, 5).value.replace("tel_1",tel)
                    elif sheet.cell(case_id+1,5).value.find("tel")!=-1:
                        row_data['data'] = sheet.cell(case_id + 1, 5).value.replace("tel", tel+str(int(2)))
                    else:
                        row_data["data"] = sheet.cell(case_id+1, 5).value
                    row_data["expected"] = sheet.cell(case_id+1, 6).value
                    row_data["method"] = sheet.cell(case_id+1, 7).value
                    row_data["sheet_name"] = key  # 加一个表单名
                    test_data.append(row_data)

        return test_data

    def write_back(self,file_name,sheet_name,i,result,TestResult): #专门写回数据
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,8).value=result
        sheet.cell(i, 9).value = TestResult
        wb.save(file_name)  #保存结果

if __name__ == '__main__':
    res=Do_Excel().get_data(project_path.test_case_path)
    print(res)